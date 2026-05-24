from __future__ import annotations

import os
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from src.core.csv_utils import sanitize_csv_cell


def sanitize_xlsx_cell(value: Any) -> Any:
    value = sanitize_csv_cell(value)
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


class XlsxRowWriter:
    def __init__(
        self,
        output_path: str,
        fieldnames: Iterable[str],
        sheet_name: str = "数据",
        autosave_every: int = 1,
    ):
        self.output_path = str(output_path)
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        self.fieldnames = list(fieldnames)
        self.autosave_every = max(1, int(autosave_every or 1))
        self._rows_since_save = 0
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name[:31] or "数据"
        self.worksheet.append(self.fieldnames)
        self.save()

    def writerow(self, row: Mapping[str, Any]):
        self._append_row(row)
        self._rows_since_save += 1
        if self._rows_since_save >= self.autosave_every:
            self.save()

    def writerows(self, rows: Iterable[Mapping[str, Any]]):
        wrote_rows = False
        for row in rows:
            self._append_row(row)
            wrote_rows = True
        if wrote_rows:
            self.save()

    def _append_row(self, row: Mapping[str, Any]):
        self.worksheet.append([sanitize_xlsx_cell(row.get(field, "")) for field in self.fieldnames])

    def save(self):
        temp_path = f"{self.output_path}.tmp"
        self.workbook.save(temp_path)
        try:
            os.replace(temp_path, self.output_path)
        except OSError:
            self.workbook.save(self.output_path)
        self._rows_since_save = 0



class MultiSheetXlsxWriter:
    def __init__(
        self,
        output_path: str,
        sheets_fields: dict[str, list[str]],
        autosave_every: int = 1,
    ):
        self.output_path = str(output_path)
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        self.sheets_fields = sheets_fields
        self.autosave_every = max(1, int(autosave_every or 1))
        self._rows_since_save = 0
        
        self.workbook = Workbook()
        # Remove default sheet created by openpyxl
        default_sheet = self.workbook.active
        if default_sheet is not None:
            self.workbook.remove(default_sheet)
            
        self.worksheets = {}
        for sheet_name, fieldnames in sheets_fields.items():
            ws = self.workbook.create_sheet(title=sheet_name[:31] or "Sheet")
            ws.append(list(fieldnames))
            self.worksheets[sheet_name] = ws
        self.save()

    def writerow(self, sheet_name: str, row: Mapping[str, Any]):
        if sheet_name not in self.worksheets:
            import logging
            logging.getLogger(__name__).warning("writerow: sheet '%s' not registered, row skipped", sheet_name)
            return
        fieldnames = self.sheets_fields[sheet_name]
        ws = self.worksheets[sheet_name]
        ws.append([sanitize_xlsx_cell(row.get(field, "")) for field in fieldnames])
        self._rows_since_save += 1
        if self._rows_since_save >= self.autosave_every:
            self.save()

    def save(self):
        temp_path = f"{self.output_path}.tmp"
        self.workbook.save(temp_path)
        try:
            os.replace(temp_path, self.output_path)
        except OSError:
            self.workbook.save(self.output_path)
        self._rows_since_save = 0

