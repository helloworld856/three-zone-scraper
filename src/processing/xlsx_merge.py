from __future__ import annotations

import argparse
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.core import build_output_path, sanitize_xlsx_cell


PLATFORM_PREFIX = {
    "youtube": "youtube",
    "tiktok": "tiktok",
    "x": "x",
    "twitter": "x",
}


def normalize_platform(platform: str) -> str:
    value = (platform or "").strip().lower()
    if value in PLATFORM_PREFIX:
        return PLATFORM_PREFIX[value]
    return value or "merged"


def find_xlsx_files(folder: str | Path, keyword: str, output_file: str | Path | None = None) -> list[Path]:
    folder_path = Path(folder)
    keyword = (keyword or "").strip().lower()
    output_name = Path(output_file).name.lower() if output_file else ""
    files: list[Path] = []
    for path in sorted(folder_path.glob("*.xlsx")):
        name = path.name.lower()
        if output_name and name == output_name:
            continue
        if path.name.startswith("~$"):
            continue
        if keyword and keyword not in name:
            continue
        files.append(path)
    return files


def _normalize_headers(raw_headers) -> list[str]:
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    while headers and not headers[-1]:
        headers.pop()
    return headers


def merge_xlsx_files(
    folder: str | Path,
    keyword: str = "keyword",
    platform: str = "merged",
    output_file: str | Path | None = None,
) -> tuple[str, int, int]:
    platform_prefix = normalize_platform(platform)
    output_path = Path(output_file) if output_file else Path(
        build_output_path(platform_prefix, f"{platform_prefix}_merge_{time.strftime('%Y%m%d')}.xlsx")
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    files = find_xlsx_files(folder, keyword, output_path)
    if not files:
        if keyword:
            raise FileNotFoundError(f"没有找到文件名包含“{keyword}”的 .xlsx 文件")
        raise FileNotFoundError("没有找到可合并的 .xlsx 文件")

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "合并数据"

    headers: list[str] | None = None
    serial_col_index: int | None = None
    current_no = 1
    merged_rows = 0
    merged_files = 0

    for file_path in files:
        wb = None
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            file_row_count = 0

            for ws in wb.worksheets:
                row_iter = ws.iter_rows(values_only=True)
                source_headers = _normalize_headers(next(row_iter, []))
                if not source_headers or all(not value for value in source_headers):
                    continue

                if headers is None:
                    headers = list(source_headers)
                    if "序号" in headers:
                        serial_col_index = headers.index("序号")
                    else:
                        headers.insert(0, "序号")
                        serial_col_index = 0
                    output_ws.append(headers)

                source_index = {name: index for index, name in enumerate(source_headers)}
                for row_values in row_iter:
                    if not row_values or all(value is None or str(value).strip() == "" for value in row_values):
                        continue
                    output_row = []
                    for column_index, header in enumerate(headers):
                        if column_index == serial_col_index:
                            output_row.append(current_no)
                        else:
                            source_pos = source_index.get(header)
                            value = row_values[source_pos] if source_pos is not None and source_pos < len(row_values) else ""
                            output_row.append(sanitize_xlsx_cell(value))
                    output_ws.append(output_row)
                    current_no += 1
                    merged_rows += 1
                    file_row_count += 1

            if file_row_count:
                merged_files += 1
                output_wb.save(output_path)
        except Exception as exc:
            print(f"警告：跳过文件 {file_path.name}（{exc}）")
        finally:
            if wb is not None:
                wb.close()

    if merged_rows <= 0:
        raise ValueError("没有成功读取任何有效数据，请检查文件、关键词或表头")

    output_wb.save(output_path)
    return str(output_path), merged_files, merged_rows


def main(argv=None):
    parser = argparse.ArgumentParser(description="合并多个 xlsx 文件")
    parser.add_argument("folder", help="包含 xlsx 文件的文件夹")
    parser.add_argument("--keyword", default="keyword", help="文件名包含的关键词，留空则合并所有 xlsx")
    parser.add_argument("--platform", default="merged", help="平台前缀，例如 youtube/tiktok/x")
    parser.add_argument("--output", default="", help="输出 xlsx 路径，不填则自动写入 output/<platform>")
    args = parser.parse_args(argv)
    merge_xlsx_files(args.folder, args.keyword, args.platform, args.output or None)


if __name__ == "__main__":
    main()
